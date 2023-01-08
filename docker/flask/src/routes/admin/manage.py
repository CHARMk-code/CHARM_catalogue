from flask import Blueprint, send_from_directory, request, send_file, session
from werkzeug.utils import secure_filename
from flask_cors import CORS
import os,sys, datetime, math
from ...models import Company, Layout, Prepage,  Tag, Map
from flask_api import status
from ... import db, config
from ...helper_functions import *
import shutil
import openpyxl
import urllib.parse


ACCEPT_IMAGE_EXTENDS = ["jpg","png","svg"] 
NUMBER_OF_METADATA_COLS_COMPANY = 17
NUMBER_OF_METADATA_COLS_TAG = 6
blueprint = Blueprint('manage', __name__, url_prefix='/api/manage')
CORS(blueprint,origins="*", resources=r'*', allow_headers=[
    "Content-Type", "Authorization", "Access-Control-Allow-Credentials"])

@blueprint.route("/image/<filename>", methods = ["GET"])
def imageSend(filename):
    return send_from_directory(config['flask']['static_folder'], (filename))


def imageLoad(request):
    file = request.files["file"]
    filename = file.filename
    if not filename[-3:] in ACCEPT_IMAGE_EXTENDS:
        return f'{filename} is not accept file type', status.HTTP_400_BAD_REQUEST

    urlEncodeFilename = urllib.parse.unquote_plus(filename)
    file.save(os.path.join(config['flask']['static_folder'], filename))
    return "All files uploaded", status.HTTP_200_OK


def parseXlsx():
    # Inactives company
    Company.query.update({Company.active:False})
    db.session.commit()

    workbook = openpyxl.load_workbook(os.path.join(config["flask"]["upload_folder"],"CHARM_CATALOGUE_DATA.xlsx"))

    # Adds maps
    maps_sheet = workbook["Maps"]

    for i in range(2,maps_sheet.max_row + 1):
        map_object = Prepage.query.filter_by(name=maps_sheet.cell(i,1)).first()

        data = list(map(lambda x: x.value, maps_sheet[i]))
        REF_NAME_POS = 2
        data[REF_NAME_POS] = mapLookUpIdOrNull(data[REF_NAME_POS])
        if not map_object:
            Map.create(*data)
        else:
            map_object.update(*data)

    # Prepages
    prepages_sheet = workbook["Prepages"]
    for i in range(2,prepages_sheet.max_row + 1):
        prepage = Prepage.query.filter_by(name=prepages_sheet.cell(i,1)).first()

        data = list(map(lambda x: x.value, prepages_sheet[i]))
        if not prepage:
            Prepage.create(*data)
        else:
            prepage.update(*data)

    # Layout
    layout_sheet = workbook["Layout"]
    for i in range(2,layout_sheet.max_row + 1):
        layout = Layout.query.filter_by(image=layout_sheet.cell(i,2)).first()

        data = list(map(lambda x: x.value, layout_sheet[i]))
        if not layout:
            Layout.create(*data)
        else:
            layout.update(*data)


    # Adds tags
    tags_sheet = workbook["Tags"]
    for i in range(2,tags_sheet.max_row + 1):
        row = tags_sheet[i]
        tag = Tag.query.filter_by(name=row[NUMBER_OF_METADATA_COLS_TAG].value).first()
        metadata = list(map( lambda x: x.value, row[:NUMBER_OF_METADATA_COLS_TAG]))
        if not tag: # No tag exists
            Tag.create(row[NUMBER_OF_METADATA_COLS_TAG].value,None,1,1,False,*metadata)
        else:
            tag.update(row[NUMBER_OF_METADATA_COLS_TAG].value,None,1,1,False,*metadata)


   # next_col = NUMBER_OF_METADATA_COLS_TAG
   # parent_tag = None
   # for i in range(2,tags_sheet.max_row + 1):
   #     row = tags_sheet[i]
   #     tag = Tag.query.filter_by(name=row[next_col].value).first()
   #     metadata = list(map( lambda x: x.value, row[:NUMBER_OF_METADATA_COLS_TAG]))
   #     if not tag: # No tag exists
   #         Tag.create(row[next_col].value,parent_tag,1,1,False,*metadata)
   #         parent_tag = Tag.query.filter_by(name=row[next_col].value).first().id
   #     else:
   #         tag.update(row[next_col].value,parent_tag,1,1,False,*metadata)
   #         parent_tag = tag.id
   #
   #     if (i+1 >= tags_sheet.max_row):
   #         break
   #     if (tags_sheet.cell(i+1,next_col)!=''):
   #         if (next_col==NUMBER_OF_METADATA_COLS_TAG):
   #             parent_tag = None
   #         continue
   #     elif (next_col+1 < tags_sheet.ncols):
   #         if (tags_sheet.cell(i+1,next_col+1)!=''):
   #             next_col += 1
   #             continue
##
   #     print(tags_sheet, file=sys.stderr)
   #     for j in range(tags_sheet.ncols):
   #         if tags_sheet.cell(i+1,j) != '':
   #             if (j==NUMBER_OF_METADATA_COLS_TAG):
   #                 parent_tag = None
   #             next_col = j
   #             break
   #         if (tags_sheet.cell(i+1,next_col)!=''):
   ##             if (next_col==NUMBER_OF_METADATA_COLS_TAG):
   #                 parent_tag = None
   #             continue
   #         elif (next_col+1 < tags_sheet.ncols):
  # #             if (tags_sheet.cell(i+1,next_col+1)!=''):
  #                  next_col += 1
   #                 continue

     #       for j in range(tags_sheet.ncols):
     #           if tags_sheet.cell(i+1,j) != '':
     #               if (j==NUMBER_OF_METADATA_COLS_TAG):
     #                   parent_tag = None
     #               next_col = j
      #              break
    # Generats companies
    companies_sheet = workbook["Companies"]
    tags = []

    tag_row = companies_sheet[1]
    with db.session.no_autoflush:
        for i in range(NUMBER_OF_METADATA_COLS_COMPANY,companies_sheet.max_column):
            tags.append(Tag.query.filter_by(name = tag_row[i].value).first())

        for i in range(2,companies_sheet.max_row ):
            tags_temp = []
            for j in range(NUMBER_OF_METADATA_COLS_COMPANY +1 ,companies_sheet.max_column):
                if companies_sheet.cell(i,j):
                    tags_temp.append(tags[j-NUMBER_OF_METADATA_COLS_COMPANY])

            metadata = companies_sheet[i][:NUMBER_OF_METADATA_COLS_COMPANY]
            metadata = list(map(lambda x: x.value, metadata))

            # Map map name to id
            MAP_POS = 15

            metadata[MAP_POS] = mapLookUpIdOrNull(metadata[MAP_POS])

            if metadata[0] == "":
                continue
            print(metadata, file=sys.stderr)
            company = Company.query.filter_by(name = metadata[0]).first()
            if  company == None:
                Company.create(
                        *metadata,
                        tags_temp
                        )
            else:
                company.update(*metadata,tags_temp)

    db.session.flush()


    os.remove(os.path.join(config["flask"]["upload_folder"],"CHARM_CATALOGUE_DATA.xlsx"))

def mapLookUpIdOrNull(name):
    map_obj = Map.query.filter_by(name=name).first()
    if map_obj:
        return map_obj.id
    else:
        return None


def unpackAndParse(request):
    file = request.files["file"]
    filename = file.filename
    file_extension = filename[filename.index("."):]
    packedPath = os.path.join(config['flask']['upload_folder'], f"tmp{file_extension}" )
    file.save(packedPath)

    unpackedPath = os.path.join(config['flask']['upload_folder'], "tmp")
    shutil.unpack_archive(packedPath,unpackedPath)
    os.remove(packedPath)

    for root, dirs, files in os.walk(unpackedPath):
        for file in files:
            path = os.path.join(root,file)
            if ".xlsx" in file:
                os.rename(path, os.path.join(config["flask"]["upload_folder"],"CHARM_CATALOGUE_DATA.xlsx"))
                parseXlsx()
            elif any(map(lambda x: x in file, ACCEPT_IMAGE_EXTENDS)):
                shutil.move(path,os.path.join(config['flask']['static_folder'],file))
    shutil.rmtree(unpackedPath)
    return "", status.HTTP_200_OK

@blueprint.route("/upload", methods=["POST"])
def load():
    """
    POST endpoint api/manage/load

    Allow for for uploading of images and data, see README
    """
    result = auth_token(request)
    if not result[0]:
        return result[1]
    if "file" not in request.files:
        return "No file found",status.HTTP_400_BAD_REQUEST

    file = request.files["file"]

    if ".xlsx" in file.filename: # load data
        file.save(os.path.join(config["flask"]["upload_folder"],"CHARM_CATALOGUE_DATA.xlsx"))
        parseXlsx()
    elif any(map(lambda x: x in file.filename, ACCEPT_IMAGE_EXTENDS)): # Load single image
        imageLoad(request)
    elif any(map(lambda x:x in file.filename, [".zip", ".tar.gz"])):
            unpackAndParse(request)
    else:
        return f"{file.filename} unaccepted file", status.HTTP_400_BAD_REQUEST
    return "Success", status.HTTP_201_CREATED



@blueprint.route("/download", methods=["GET"])
def download():
    result = auth_token(request)
    if not result[0]:
        return result[1]

    data_dump_name = f"CHARM_catalogue_datadump_{datetime.datetime.now().date()}"
    data_dump_path = f"/tmp/{data_dump_name}"

    # Copy all static resources into temporary folder
    if os.path.exists(data_dump_path):
        shutil.rmtree(data_dump_path)
    shutil.copytree(config["flask"]["static_folder"], data_dump_path)

    # Sort into images into directory to simplify for user

    # (directory name, Table name, column name)
    tables = [("Logos", Company, "logo"), ("Prepages", Prepage, "image"), ("Tag_icons", Tag, "icon"), ("Map", Map, "image"), ("Layouts", Layout, "image")]
    for (name, table, col) in tables:
        os.makedirs(data_dump_path+f"/{name}")
        for row in table.query.all():
            image_file = getattr(row,col)
            if image_file != "" and os.path.exists(f"{data_dump_path}/{image_file}"):
                shutil.move(f"{data_dump_path}/{image_file}",data_dump_path+f"/{name}")


    # Generate xlsx from db
    workbook  = xlsxwriter.Workbook(f'{data_dump_path}/CHARM_CATALOGUE_DATA.xlsx')

    # Populate simple sheets, as there have simulater structure we can generalize them
    metadata= [
        ("Maps", Map, ["Name","Image","Ref"],["name","image","ref"]),
        ("Prepages",Prepage,["Name","Active","Image","Order"],["name","active","image","order"]),
        ("Layout",Layout,["Active","Image","PLACMENT (0 = random center, 1 = left sidebar, 2 = right sidebar)"],
        ["active","image","placement"]),
        ("Tags",Tag,["Logo","Divsion","Business Area","Looking for","Offering","Language","Name"],
        ["icon","division","business_area","looking_for","offering","language","name"])
    ]
    for (sheet_name, table,labels,attrs) in metadata:
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.write_row(0,0, labels)


        row_num = 1
        for table_obj in table.query.all():
            row_data = [getattr(table_obj,attr) for attr in attrs]
            if sheet_name == "Maps" and row_data[2] != None:
                row_data[2] = Map.query.filter_by(id=row_data[2]).first().name
            worksheet.write_row(row_num,0, row_data)
            row_num += 1
    

    # Special case for companies
    worksheet = workbook.add_worksheet("Companies")
    
    # Set object based labels
    labels=["Name","Active","CHARMTALK","Summer job description", "Summer job link","Description","Contact","Contact email", "Employees worldwide","Website","Talk to us about", "Logo","Map","Booth number"]
    worksheet.write_row(0,0, labels)

    # Set tag based labels
    all_tags = Tag.query.all()
    name_for_all_tags = list(map(lambda x: x.name, all_tags))
    id_for_all_tags = list(map(lambda x: x.id, all_tags))
    worksheet.write_row(0,NUMBER_OF_METADATA_COLS_COMPANY,name_for_all_tags)

    row_num = 1
    attrs= ["name","active","charmtalk","summer_job_description", "summer_job_link","description","contacts","contact_email","employees_world","website","talk_to_us_about","logo","map_image","booth_number"]
    # Tag data
    for table_obj in Company.query.all():
        # Object based data
        row_data = [getattr(table_obj,attr) for attr in attrs]
        worksheet.write_row(row_num,0,row_data)

        # Tag data
        offset = 0
        company_tag_ids = list(map(lambda x:x.id, table_obj.tags))
        for id_to_test in id_for_all_tags:
            worksheet.write(row_num,NUMBER_OF_METADATA_COLS_COMPANY+offset, id_to_test in company_tag_ids)
            offset +=1
        
        row_num += 1
    

    workbook.close()
    # Pack and send result
    shutil.make_archive(data_dump_path,"zip",data_dump_path)
    shutil.rmtree(data_dump_path)
    return send_file(data_dump_path+".zip")
